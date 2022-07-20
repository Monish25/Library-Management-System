import os
from tkinter import messagebox
from tkinter import *
from openpyxl import *
#openpyxl call
data = load_workbook('E:\Library Management\database.xlsx')
sheet = data.active


def excel():
    #writing data for the file
    
    sheet.cell(row=1, column=1).value = "Srl.No"
    sheet.cell(row=1, column=2).value = "Book Group"
    sheet.cell(row=1, column=3).value = "Serial Number"
    sheet.cell(row=1, column=4).value = "Book Title"
    sheet.cell(row=1, column=5).value = "Publication Country"
    sheet.cell(row=1, column=6).value = "Publication State"
    sheet.cell(row=1, column=7).value = "User Access"
    sheet.cell(row=1, column=8).value = "Book Author"
    sheet.cell(row=1, column=9).value = "Publisher"
    sheet.cell(row=1, column=10).value = "No of Pages"
    sheet.cell(row=1, column=11).value = "Vendor"
    sheet.cell(row=1, column=12).value = "Language"
    sheet.cell(row=1, column=13).value = "Price"


# the focus sets the textbox to the line
def focus1(event):
    srl_no_field.focus_set()
    
def focus2(event):
    bg_field.focus_set()
    
def focus3(event):
    s_no_field.focus_set()
    
def focus4(event):
    book_title_field.focus_set()
    
def focus5(event):
    pub_ctry_field.focus_set()
    
def focus6(event):
    pub_st_field.focus_set()
    
def focus7(event):
    ua_field.focus_set()
    
def focus8(event):
    book_ath_field.focus_set()
    
def focus9(event):
    pub_field.focus_set()
    
def focus10(event):
    pgno_field.focus_set()
    
def focus11(event):
    vendor_field.focus_set()
    
def focus12(event):
    lang_field.focus_set()

def focus13(event):
    price_field.focus_set()
    
    
def clear():   # clear the value once data is entered
    srl_no_field.delete(0,END)
    bg_field.delete(0,END)
    s_no_field.delete(0,END)
    book_title_field.delete(0,END)
    pub_ctry_field.delete(0,END)
    pub_st_field.delete(0,END)
    ua_field.delete(0,END)
    book_ath_field.delete(0,END)
    pub_field.delete(0,END)
    pgno_field.delete(0,END)
    vendor_field.delete(0,END)
    lang_field.delete(0,END)
    price_field.delete(0,END)
    

def insert():
    if (srl_no_field.get() == "" and
        bg_field.get() == "" and
        s_no_field.get() == "" and
        book_title_field.get() == "" and
        pub_ctry_field.get() == "" and
        pub_st_field.get() == "" and
        ua_field.get() == "" and
        book_ath_field.get() == "" and
        pub_field.get() == "" and
        pgno_field.get() == "" and
        vendor_field.get() == "" and
        lang_field.get() == "" and
        price_field.get() == "" ):
        
        messagebox.showinfo('No Value', 'Enter a Value' ) ### when user did not enter value
    else:
        current_row = sheet.max_row 
        current_column = sheet.max_column
    #finally adds the data
        sheet.cell(row=current_row + 1, column=1).value = srl_no_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = bg_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = s_no_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = book_title_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = pub_ctry_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = pub_st_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = ua_field.get() 
        sheet.cell(row=current_row + 1, column=8).value = book_ath_field.get()
        sheet.cell(row=current_row + 1, column=9).value = pub_field.get()
        sheet.cell(row=current_row + 1, column=10).value = pgno_field.get()
        sheet.cell(row=current_row + 1, column=11).value = vendor_field.get()
        sheet.cell(row=current_row + 1, column=12).value = lang_field.get()
        sheet.cell(row=current_row + 1, column=13).value = price_field.get()
        messagebox.showinfo('Entered', 'The data Has been Succesfully Entered')
        
        data.save('E:\Library Management\database.xlsx')   ##saving the file
        
        srl_no_field.focus_set()   #setting the focus back on the name field
        #clearing the data
        clear()

#def launching the file
def file():
    os.system('database.xlsx')
###starting the defintion of the GUI  
a = Tk()
a.title('Insert Books')
a.geometry('500x400')     ##define the tkinter window
#labels start
heading = Label(a, text = 'Insert Books',font = ("Lucida Handwriting", 13))
lab1 = Label(a, text = 'srl.no')
lab2 = Label(a, text = 'book group')
lab3 = Label(a, text = 'series')
lab4 = Label(a, text = 'book title')
lab5 = Label(a, text = 'publication country')
lab6 = Label(a, text = 'publication state')
lab7 = Label(a, text = 'user access')
lab8 = Label(a, text = 'book author')
lab9 = Label(a, text = 'Publisher')
lab10 = Label(a, text = 'no of pages')
lab11 = Label(a, text = 'vendor')
lab12 = Label(a, text = 'language')
lab13 = Label(a, text = 'price')


## label positioning and activation

heading.grid(row=0, column=1) 
lab1.grid(row=1, column=0) 
lab2.grid(row=2, column=0) 
lab3.grid(row=3, column=0) 
lab4.grid(row=4, column=0) 
lab5.grid(row=5, column=0) 
lab6.grid(row=6, column=0) 
lab7.grid(row=7, column=0) 
lab8.grid(row=8, column=0) 
lab9.grid(row=9, column=0) 
lab10.grid(row=10, column=0) 
lab11.grid(row=11, column=0) 
lab12.grid(row=12, column=0) 
lab13.grid(row=13, column=0) 
 

## textbox setup 
srl_no_field = Entry(a)
bg_field = Entry(a)
s_no_field = Entry(a)
book_title_field = Entry(a)
pub_ctry_field = Entry(a)
pub_st_field = Entry(a)
ua_field = Entry(a)
book_ath_field = Entry(a)
pub_field = Entry(a)
pgno_field = Entry(a)
vendor_field = Entry(a)
lang_field = Entry(a)
price_field = Entry(a)


#calling focus buttons for respective textboxes when enter is pressed
srl_no_field.bind("<Return>", focus2)
bg_field.bind("<Return>", focus3)
s_no_field.bind("<Return>", focus4)
book_title_field.bind("<Return>", focus5)
pub_ctry_field.bind("<Return>", focus6)
pub_st_field.bind("<Return>", focus7)
ua_field.bind("<Return>", focus8)
book_ath_field.bind("<Return>", focus9)
pub_field.bind("<Return>", focus10)
pgno_field.bind("<Return>", focus11)
vendor_field.bind("<Return>", focus12)
lang_field.bind("<Return>", focus13)
price_field.bind("<Return>", focus1)


#position of textboxes
srl_no_field.grid(row=1, column=1, ipadx="100")
bg_field.grid(row=2, column=1, ipadx="100")
s_no_field.grid(row=3, column=1, ipadx="100")
book_title_field.grid(row=4, column=1, ipadx="100")
pub_ctry_field.grid(row=5, column=1, ipadx="100")
pub_st_field.grid(row=6, column=1, ipadx="100")
ua_field.grid(row=7, column=1, ipadx="100")
book_ath_field.grid(row=8, column=1, ipadx="100")
pub_field.grid(row=9, column=1, ipadx="100")
pgno_field.grid(row=10, column=1, ipadx="100")
vendor_field.grid(row=11, column=1, ipadx="100")
lang_field.grid(row=12, column=1, ipadx="100")
price_field.grid(row=13, column=1, ipadx="100")



#calling the defined excel function
excel()
save = Button(a,text = 'Save', width = 4,command = insert)
save.grid(row=15, column=1)
view = Button(a, text = 'View', width = 4, command = file)
view.grid(row = 16, column = 1)
end = Button(a, text = 'Exit', width = 4,command = a.destroy)
end.grid(row = 17,column = 1)
a.mainloop()