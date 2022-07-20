import os
from tkinter import messagebox
from tkinter import *
from openpyxl import *
#openpyxl call
data = load_workbook('E:\Library Management\issue.xlsx')
sheet = data.active


def excel():
    #writing data for the file
    
    sheet.cell(row=1, column=1).value = "Invoice No"
    sheet.cell(row=1, column=2).value = "Date of Borrowing"
    sheet.cell(row=1, column=3).value = "Name of Borrowee"
    sheet.cell(row=1, column=4).value = "Contact No"
    sheet.cell(row=1, column=5).value = "Occupation"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Name Of the book"
    sheet.cell(row=1, column=8).value = "Book S.no"
    sheet.cell(row=1, column=9).value = "Book Publisher"
    sheet.cell(row=1, column=10).value = "Book Group"
    sheet.cell(row=1, column=11).value = "Pricing"
    sheet.cell(row=1, column=12).value = "Due Date"
    sheet.cell(row=1, column=13).value = "Status"


# the focus sets the textbox to the line
def focus1(event):
    inv_no_field.focus_set()
    
def focus2(event):
    dob_field.focus_set()
    
def focus3(event):
    nob_field.focus_set()
    
def focus4(event):
    cont_no_field.focus_set()
    
def focus5(event):
    occ_field.focus_set()
    
def focus6(event):
    e_id_field.focus_set()
    
def focus7(event):
    book_nm_field.focus_set()
    
def focus8(event):
    book_sno_field.focus_set()
    
def focus9(event):
    book_pub_field.focus_set()
    
def focus10(event):
    book_grp_field.focus_set()
    
def focus11(event):
    price_field.focus_set()
    
def focus12(event):
    dd_field.focus_set()

def focus13(event):
    status_field.focus_set()
    
    
def clear():   # clear the value once data is entered
    inv_no_field.delete(0,END)
    dob_field.delete(0,END)
    nob_field.delete(0,END)
    cont_no_field.delete(0,END)
    occ_field.delete(0,END)
    e_id_field.delete(0,END)
    book_nm_field.delete(0,END)
    book_sno_field.delete(0,END)
    book_pub_field.delete(0,END)
    book_grp_field.delete(0,END)
    price_field.delete(0,END)
    dd_field.delete(0,END)
    status_field.delete(0,END)
    

def insert():
    if (inv_no_field.get() == "" and
        dob_field.get() == "" and
        nob_field.get() == "" and
        cont_no_field.get() == "" and
        occ_field.get() == "" and
        e_id_field.get() == "" and
        book_nm_field.get() == "" and
        book_sno_field.get() == "" and
        book_pub_field.get() == "" and
        book_grp_field.get() == "" and
        price_field.get() == "" and
        dd_field.get() == "" and
        status_field.get() == "" ):
        
        messagebox.showinfo('No Value', 'Enter a Value' ) ### when user did not enter value
    else:
        current_row = sheet.max_row 
        current_column = sheet.max_column
    #finally adds the data
        sheet.cell(row=current_row + 1, column=1).value = inv_no_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = dob_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = nob_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = cont_no_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = occ_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = e_id_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = book_nm_field.get() 
        sheet.cell(row=current_row + 1, column=8).value = book_sno_field.get()
        sheet.cell(row=current_row + 1, column=9).value = book_pub_field.get()
        sheet.cell(row=current_row + 1, column=10).value = book_grp_field.get()
        sheet.cell(row=current_row + 1, column=11).value = price_field.get()
        sheet.cell(row=current_row + 1, column=12).value = dd_field.get()
        sheet.cell(row=current_row + 1, column=13).value = status_field.get()
        messagebox.showinfo('Entered', 'The data Has been Succesfully Entered')
        
        
        data.save('E:\Library Management\issue.xlsx')   ##saving the file
        
        inv_no_field.focus_set()   #setting the focus back on the name field
        #clearing the data
        clear()

#def launching the file
def file():
    os.system('issue.xlsx')
###starting the defintion of the GUI  
a = Tk()
a.title('Issue Books')
a.geometry('500x400')     ##define the tkinter window
#labels start
heading = Label(a, text = 'Issue',font = ("Lucida Handwriting", 13))
lab1 = Label(a, text = 'Invoice Number')
lab2 = Label(a, text = 'Date of Borrowing')
lab3 = Label(a, text = 'Name of Borrowee')
lab4 = Label(a, text = 'Contact Number')
lab5 = Label(a, text = 'Occupation')
lab6 = Label(a, text = 'Email ID')
lab7 = Label(a, text = 'Name of the Book')
lab8 = Label(a, text = 'Book Serial Number')
lab9 = Label(a, text = 'Book Publisher')
lab10 = Label(a, text = 'Book Group')
lab11 = Label(a, text = 'Pricing')
lab12 = Label(a, text = 'Due Day')
lab13 = Label(a, text = 'Status')


## label positioning and activation

heading.grid(row=0, column=1) 
lab1.grid(row=1, column=0) 
lab2.grid(row=2, column=0) 
lab3.grid(row=3, column=0) 
lab4.grid(row=4, column=0) 
lab5.grid(row=5, column=0) 
lab6.grid(row=6, column=0) 
lab7.grid(row=7, column=0) 
lab8.grid(row=8, column=0) 
lab9.grid(row=9, column=0) 
lab10.grid(row=10, column=0) 
lab11.grid(row=11, column=0) 
lab12.grid(row=12, column=0) 
lab13.grid(row=13, column=0) 
 

## textbox setup 
inv_no_field = Entry(a)
dob_field = Entry(a)
nob_field = Entry(a)
cont_no_field = Entry(a)
occ_field = Entry(a)
e_id_field = Entry(a)
book_nm_field = Entry(a)
book_sno_field = Entry(a)
book_pub_field = Entry(a)
book_grp_field = Entry(a)
price_field = Entry(a)
dd_field = Entry(a)
status_field = Entry(a)


#calling focus buttons for respective textboxes when enter is pressed
inv_no_field.bind("<Return>", focus2)
dob_field.bind("<Return>", focus3)
nob_field.bind("<Return>", focus4)
cont_no_field.bind("<Return>", focus5)
occ_field.bind("<Return>", focus6)
e_id_field.bind("<Return>", focus7)
book_nm_field.bind("<Return>", focus8)
book_sno_field.bind("<Return>", focus9)
book_pub_field.bind("<Return>", focus10)
book_grp_field.bind("<Return>", focus11)
price_field.bind("<Return>", focus12)
dd_field.bind("<Return>", focus13)
status_field.bind("<Return>", focus1)


#position of textboxes
inv_no_field.grid(row=1, column=1, ipadx="100")
dob_field.grid(row=2, column=1, ipadx="100")
nob_field.grid(row=3, column=1, ipadx="100")
cont_no_field.grid(row=4, column=1, ipadx="100")
occ_field.grid(row=5, column=1, ipadx="100")
e_id_field.grid(row=6, column=1, ipadx="100")
book_nm_field.grid(row=7, column=1, ipadx="100")
book_sno_field.grid(row=8, column=1, ipadx="100")
book_pub_field.grid(row=9, column=1, ipadx="100")
book_grp_field.grid(row=10, column=1, ipadx="100")
price_field.grid(row=11, column=1, ipadx="100")
dd_field.grid(row=12, column=1, ipadx="100")
status_field.grid(row=13, column=1, ipadx="100")



#calling the defined excel function
excel()
save = Button(a,text = 'Save', width = 4,command = insert)
save.grid(row=15, column=1)
view = Button(a, text = 'View', width = 4, command = file)
view.grid(row = 16, column = 1)
end = Button(a, text = 'Exit', width = 4,command = a.destroy)
end.grid(row = 17,column = 1)
a.mainloop()