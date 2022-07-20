from tkinter import messagebox
from tkinter import *
import openpyxl
from openpyxl import *

data = load_workbook('E:\Library Management\issr.xlsx')
sheet = data.active


def update_text(info):
    book_info.delete(1.0, 'end')
    book_info.insert('end', info)

def value():
    if (inv_field.get() == '' and
        date_field.get() == ''):
        messagebox.showinfo('NO Value','There is No value inserted')
    else:
        find_book()

def find_book():
    inv_no = inv_field.get()
    if inv_no:
            wb = openpyxl.load_workbook('E:\Library Management\issue.xlsx')
            sheet = wb.active
            for row in sheet.rows:
                # assume invoice no is in column 1
                if row[0].value == inv_no:
                    update_text('\n'.join(str(cell.value) if cell.value else '' for cell in row))
                    return
            wb.close()
            update_text('Issue Entry not found')


def date():
    if date_field.get() == '':
        messagebox.showinfo('No Date','There is No date Entered')
    else:
        save()


       
def save():
    inv_no = inv_field.get()
    date_fetch = date_field.get()
    
    current_row = sheet.max_row 
    current_column = sheet.max_column
    sheet.cell(row=current_row + 1, column=1).value = inv_no
    sheet.cell(row=current_row + 1, column=2).value = date_fetch
    sheet.cell(row=1, column=1).value = "Invoice No"
    sheet.cell(row=1, column=2).value = "Date of Returning"
    data.save('E:\Library Management\issr.xlsx')
    messagebox.showinfo('Saved','The Data Has Been Saved')
    clear()
    
def clear():
    inv_field.delete(0,END)
    date_field.delete(0,END)
    
    
    
a = Tk()
a.title('Return Book')
a.geometry('500x230')
heading = Label(a,text = 'Return Book')
heading.grid(row = 0,column = 1)
lab1 = Label(a,text = 'Enter Invoice Number:')
lab1.grid(row = 1, column = 0)
inv_field = Entry(a)
inv_field.grid(row = 1, column = 1)
inv_field.get()
find = Button(a,text = 'Find',width = 4,command =value)
find.grid(row = 2, column = 1)
book_info = Text(a, width=40, height=5)
book_info.grid(row = 3 ,column = 1)
savebut = Button(a,text = 'Save',width = 4,command = date)
savebut.grid(row = 5,column = 1)
datelab = Label(a,text = 'Enter The Date')
datelab.grid(row =4,column = 0)
date_field = Entry(a)
date_field.grid(row = 4,column = 1)
ext = Button(a,text = 'Exit',width = 4, command = a.destroy)
ext.grid(row = 6, column = 1)

a.mainloop()